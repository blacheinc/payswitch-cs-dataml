"""
Excel format detector
Detects Excel structure: sheets, headers, data ranges
"""

import io
from typing import Dict, Any, Optional, Tuple, List

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from system_interfaces import FileIntrospectionResult
except ImportError:
    from ..system_interfaces import FileIntrospectionResult

from .base import BaseFormatDetector
from .common import collect_evidence


class ExcelDetector(BaseFormatDetector):
    """
    Excel format detector
    Detects Excel structure with sheet detection and header analysis
    """
    
    # Maximum file size to read (10MB default, can be overridden)
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    
    def detect_format_signature(
        self,
        sample_data: bytes,
        introspection_result: Optional[FileIntrospectionResult] = None
    ) -> Tuple[bool, float, List[str]]:
        """
        Detect if data matches Excel format
        
        Args:
            sample_data: Sample bytes from file
            introspection_result: Result from File Introspection System
            
        Returns:
            Tuple of (is_match, confidence, evidence)
        """
        evidence = []
        confidence = 0.0
        
        # Check format hints from introspection
        if introspection_result and introspection_result.format_hints:
            if introspection_result.format_hints.get('possible_excel'):
                confidence += 0.3
                evidence.append("Excel format hints detected in introspection")
        
        # Check for Excel magic bytes (ZIP-based format)
        excel_magic_bytes = [b'PK\x03\x04', b'PK\x05\x06', b'PK\x07\x08']
        for magic in excel_magic_bytes:
            if sample_data.startswith(magic):
                confidence += 0.5
                evidence.append(collect_evidence(
                    f"Excel magic bytes detected",
                    byte_offset=0,
                    additional_info={'magic_bytes': magic.hex()}
                ))
                break
        
        # Check container type (Excel files are ZIP-based)
        if introspection_result and introspection_result.container_type == 'zip':
            confidence += 0.2
            evidence.append("ZIP container detected (Excel files are ZIP-based)")
        
        # Check file extension
        if introspection_result and hasattr(introspection_result, 'file_path'):
            file_path = introspection_result.file_path
            if file_path and (file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xls')):
                confidence += 0.2
                evidence.append(f"File extension indicates Excel format")
        
        is_match = confidence >= 0.5
        return is_match, min(confidence, 1.0), evidence
    
    def parse_structure(
        self,
        file_path: str,
        sample_data: bytes,
        introspection_result: Optional[FileIntrospectionResult] = None
    ) -> Tuple[Dict[str, Any], Optional[str]]:
        """
        Parse Excel structure
        
        Args:
            file_path: Path to Excel file in Data Lake
            sample_data: Sample bytes from file (not used, we need full file)
            introspection_result: Result from File Introspection System
            
        Returns:
            Tuple of (structure_info, error)
        """
        evidence = []
        metadata = {}
        
        if not EXCEL_AVAILABLE:
            error = "openpyxl not available, cannot parse Excel structure"
            evidence.append(collect_evidence(error))
            return {}, error
        
        # Get file size first
        try:
            file_client = self.datalake_client.get_file_client(file_path)
            file_properties = file_client.get_file_properties()
            file_size = file_properties.size
        except Exception as e:
            # Fallback to blob client
            if self.blob_client:
                try:
                    parts = file_path.split('/', 1)
                    if len(parts) == 2:
                        container_name, blob_path = parts
                    else:
                        container_name = None
                        blob_path = file_path
                    blob_client_instance = self.blob_client.get_blob_client(
                        container=container_name,
                        blob=blob_path
                    )
                    blob_properties = blob_client_instance.get_blob_properties()
                    file_size = blob_properties.size
                except Exception:
                    return {}, f"Failed to get file size: {str(e)}"
            else:
                return {}, f"Failed to get file size: {str(e)}"
        
        # Check file size limit
        if file_size > self.MAX_FILE_SIZE:
            error = f"Excel file size ({file_size} bytes) exceeds maximum ({self.MAX_FILE_SIZE} bytes)"
            evidence.append(collect_evidence(
                error,
                additional_info={'file_size': file_size, 'max_size': self.MAX_FILE_SIZE}
            ))
            return {}, error
        
        # Read entire Excel file (Excel files are typically small)
        excel_data, read_error = self.read_file_sample(file_path, sample_size=file_size, offset=0)
        if read_error:
            return {}, f"Failed to read Excel file: {read_error}"
        
        # Load workbook
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(excel_data), read_only=True, data_only=True)
            sheets = workbook.sheetnames
            
            evidence.append(collect_evidence(
                f"Found {len(sheets)} sheet(s)",
                additional_info={'sheets': sheets}
            ))
            
            # Analyze first sheet (typically the main data)
            if sheets:
                sheet_info, sheet_error = self._analyze_sheet(workbook[sheets[0]], sheets[0])
                if sheet_error:
                    evidence.append(collect_evidence(
                        f"Error analyzing sheet '{sheets[0]}': {sheet_error}",
                        additional_info={'sheet_name': sheets[0]}
                    ))
                    # Continue with partial results
                
                column_names = sheet_info.get('column_names')
                column_count = sheet_info.get('column_count', 0)
                row_count = sheet_info.get('row_count', 0)
                has_header = sheet_info.get('has_header', True)
                column_types = sheet_info.get('column_types', {})
                header_row = sheet_info.get('header_row')
                
                evidence.extend(sheet_info.get('evidence', []))
                
                metadata.update({
                    'sheet_count': len(sheets),
                    'active_sheet': sheets[0],
                    'header_row': header_row
                })
            else:
                # No sheets found
                column_names = None
                column_count = 0
                row_count = 0
                has_header = False
                column_types = {}
                evidence.append(collect_evidence("No sheets found in workbook"))
            
        except Exception as e:
            error = f"Excel parsing error: {str(e)}"
            evidence.append(collect_evidence(
                error,
                additional_info={'error_type': type(e).__name__}
            ))
            return {}, error
        
        return {
            'encoding': 'binary',
            'delimiter': None,
            'has_header': has_header,
            'column_count': column_count,
            'row_count': row_count,
            'nested_structure': False,
            'sheets': sheets,
            'column_names': column_names,
            'column_types': column_types,
            'evidence': evidence,
            'metadata': metadata
        }, None
    
    # ============================================================
    # Helper Methods (Excel-Specific)
    # ============================================================
    
    def _analyze_sheet(
        self,
        sheet: 'openpyxl.worksheet.worksheet.Worksheet',
        sheet_name: str
    ) -> Tuple[Dict[str, Any], Optional[str]]:
        """
        Analyze Excel sheet structure
        
        Returns:
            Tuple of (sheet_info, error)
        """
        evidence = []
        
        try:
            # Detect header row (first non-empty row)
            header_row = None
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10), start=1):
                if any(cell.value for cell in row):
                    header_row = row_idx
                    break
            
            has_header = header_row is not None and header_row == 1
            
            evidence.append(collect_evidence(
                f"Header row detected at row {header_row}" if header_row else "No header row found",
                row_number=header_row,
                additional_info={'has_header': has_header}
            ))
            
            # Get column names from header row
            column_names = None
            if has_header and header_row:
                header_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
                column_names = [
                    str(cell) if cell is not None else f"Column_{i+1}"
                    for i, cell in enumerate(header_cells)
                    if cell is not None
                ]
                evidence.append(collect_evidence(
                    f"Extracted {len(column_names)} column names from header",
                    row_number=header_row,
                    additional_info={
                        'columns': column_names[:5] if len(column_names) > 5 else column_names
                    }
                ))
            
            # Count rows and columns
            row_count = sheet.max_row
            column_count = sheet.max_column
            
            evidence.append(collect_evidence(
                f"Sheet '{sheet_name}': {row_count} rows, {column_count} columns",
                additional_info={'sheet_name': sheet_name}
            ))
            
            # Infer column types from sample data (first 100 rows)
            column_types = {}
            if column_names:
                data_rows = list(sheet.iter_rows(
                    min_row=header_row + 1 if has_header else 1,
                    max_row=min(header_row + 100 if has_header else 100, row_count),
                    values_only=True
                ))
                
                for col_idx, col_name in enumerate(column_names):
                    if col_idx < len(data_rows[0]) if data_rows else 0:
                        sample_values = [
                            row[col_idx] for row in data_rows
                            if col_idx < len(row) and row[col_idx] is not None
                        ]
                        if sample_values:
                            first_type = type(sample_values[0]).__name__
                            column_types[col_name] = first_type
                
                evidence.append(collect_evidence(
                    f"Inferred types for {len(column_types)} columns",
                    additional_info={'rows_analyzed': len(data_rows)}
                ))
            
            return {
                'column_names': column_names,
                'column_count': column_count,
                'row_count': row_count,
                'has_header': has_header,
                'header_row': header_row,
                'column_types': column_types,
                'evidence': evidence
            }, None
            
        except Exception as e:
            return {}, f"Failed to analyze sheet: {str(e)}"


# Wrapper function for backward compatibility
def detect_excel_structure(
    file_path: str,
    datalake_client,
    blob_client=None,
    introspection_result: Optional[FileIntrospectionResult] = None
) -> Dict[str, Any]:
    """
    Detect Excel structure (wrapper function for backward compatibility)
    
    Args:
        file_path: Path to Excel file in Data Lake
        datalake_client: Azure Data Lake Gen2 client
        blob_client: Optional Azure Blob Storage client for fallback
        introspection_result: Result from File Introspection System
        
    Returns:
        Dict with structure information
    """
    detector = ExcelDetector(datalake_client, blob_client)
    return detector.detect_structure(file_path, introspection_result)
